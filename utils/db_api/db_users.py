import datetime
import logging

import openpyxl


def get_users(filename='Users_2.xlsx'):
    result = []
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        result.append(ws.cell(i, 4).value)
    return result


def add_users(message, filename='Users_2.xlsx'):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    index_row = []
    unique = True
    for i in range(1, ws.max_row):
        if ws.cell(i + 1, 4).value == str(message.chat.id):
            row_user = i + 1
            unique = False
        if ws.cell(i, 1).value is None:
            index_row.append(i)

    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = list(map(lambda k: k - 1, index_row))

    if unique:
        voidrow = ws.max_row + 1
        ws.cell(column=1, row=voidrow, value=datetime.datetime.now().strftime("%d-%m-%Y %H:%M"))
        ws.cell(column=2, row=voidrow, value=message.chat.full_name)
        ws.cell(column=3, row=voidrow, value=message.chat.username)
        ws.cell(column=4, row=voidrow, value=str(message.chat.id))
        logging.info(f"В базу добавлен новый пользователь {message.chat.full_name}")
    wb.save(filename)


def check_user(message, filename='Users_2.xlsx'):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    index_row = []
    unique = True
    row_user = None
    for i in range(1, ws.max_row):
        if ws.cell(i + 1, 4).value == str(message.chat.id):
            row_user = i + 1
            unique = False
        if ws.cell(i, 1).value is None:
            index_row.append(i)

    if unique:
        return True
    else:
        if ws.cell(row_user, 5).value is None:
            return True
        else:
            return False


def info_sub(message, data, filename='Users_2.xlsx'):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    unique = True
    row_user = None
    for i in range(1, ws.max_row + 1):
        if ws.cell(i, 4).value == str(message.chat.id):
            row_user = i
            unique = False
    if unique is False:
        ws.cell(column=5, row=row_user, value=data['Q1'])
        ws.cell(column=6, row=row_user, value=data['Q2'])
        ws.cell(column=7, row=row_user, value=data['Q3'])
    wb.save(filename)
    logging.info(f"Пользователь {message.chat.full_name}({message.chat.id})"
                 f"добавлен в базу с данными: {data['Q1']}, {data['Q2']}, {data['Q3']}")





