import openpyxl


def get_category(filename='categories.xlsx'):
    result = {}
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for i in range(1, ws.max_row + 1):
        result[ws.cell(i, 1).value] = ws.cell(i, 2).value
    return result


def create_category(data, filename='categories.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    for item in data:
        ws.append([item, data[item]])
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70
    wb.save(filename)
